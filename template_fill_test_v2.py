"""テスト版 v2：Excel互換の1シート見積書テンプレートへ「値だけ」を流し込む。

このモジュールは既存の本番処理（extract_data.py / estimate_pipeline.py など）とは
独立した、テスト専用の実装です。既存ファイルや既存テンプレートには一切書き込みません。

方針:
- Numbers由来テンプレートは使わない。
- Excel互換の1シートテンプレート（シート名「見積書」）を openpyxl で新規作成する。
- 出力時は openpyxl で既存テンプレートを読み込み、指定セルへ値だけを書き込む。
- pandas.to_excel によるテンプレート全体の再生成はしない。
- 出力は最大2シート（1枚目「見積書」／2枚目「明細データ」）まで。
- 「書き出しの概要」「シート1 - 表◯」系の分解シートは作らない・許さない。
- 元テンプレートには絶対に上書きしない（別名保存＋バイト比較で検証）。
"""

from __future__ import annotations

import re
from copy import copy, deepcopy
from dataclasses import dataclass, field
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.pagebreak import Break

# ---------------------------------------------------------------------------
# 定数
# ---------------------------------------------------------------------------

BASE_DIR = Path(__file__).resolve().parent
TEMPLATE_PATH = BASE_DIR / "templates" / "test_cyca_estimate_single_sheet_template.xlsx"
OUTPUT_DIR = BASE_DIR / "output" / "template_fill_test_v2"

QUOTE_SHEET = "見積書"
DETAIL_SHEET = "明細データ"

TEMPLATE_FILL_TEST_V2_NAME = "テスト版：Excel互換テンプレート流し込み v2"

# Mac(Numbers)/Windows(Excel) 双方で崩れにくい標準フォント。
FONT_NAME = "游ゴシック"

# ブランドカラー（ロゴの紺）
BRAND_NAVY = "1F3557"
BRAND_RULE = "8A97AB"
BRAND_TINT = "EDF1F6"
LOGO_PATH = BASE_DIR / "assets" / "cyca_logo.png"

# 自社情報（正規テンプレート準拠）
COMPANY = {
    "name": "彩架建設株式会社",
    "zip": "〒839-0841",
    "address": "福岡県久留米市御井旗崎1丁目7-41",
    "tel": "TEL 0942-65-3300",
    "fax": "FAX 0942-65-3301",
    "ceo": "代表取締役　中村 哲也",
    "reg": "T3290001074771",
}

# 見積書シートの固定レイアウト（1シート・単一テーブル）
# 列: A=No. B=工事品目 C=仕様 D=数量 E=単位 F=単価 G=金額 H=備考
COL_NO, COL_ITEM, COL_SPEC, COL_QTY, COL_UNIT, COL_PRICE, COL_AMOUNT, COL_REMARK = 1, 2, 3, 4, 5, 6, 7, 8
LAST_COL = 8

TITLE_ROW = 2             # 御見積書（紺バナー）
CUSTOMER_ROW = 4          # A:C 宛名 / D 御中
GREET_ROW = 6             # A:D ごあいさつ
ISSUE_DATE_ROW = 7        # F ラベル / G:H 値
REG_ROW = 8               # F ラベル / G:H 値
AMOUNT_LABEL_ROW = 10     # A:B ラベル / C:E 御見積金額（税込）
AMOUNT_END_ROW = 11
INFO_START_ROW = 13       # 工事名称 / 工事場所 / 工事期間 / 支払条件 / 有効期限 / 見積担当
INFO_LABELS = ["工事名称", "工事場所", "工事期間", "支払条件", "有効期限", "見積担当"]
INFO_END_ROW = INFO_START_ROW + len(INFO_LABELS) - 1  # 18
COMPANY_ROW = 13          # E:H 自社情報（13〜16）

WORK_BAND_ROW = 20        # 工事内容 バンド
ITEM_HEADER_ROW = 21
ITEM_START_ROW = 22
ITEM_MAX_ROWS = 150
ITEM_END_ROW = ITEM_START_ROW + ITEM_MAX_ROWS - 1  # 171

SUMMARY_SUBTOTAL_ROW = ITEM_END_ROW + 2  # 173 小計
SUMMARY_DISCOUNT_ROW = ITEM_END_ROW + 3  # 174 値引き
SUMMARY_NET_ROW = ITEM_END_ROW + 4       # 175 税抜合計
SUMMARY_TAX_ROW = ITEM_END_ROW + 5       # 176 消費税
SUMMARY_TOTAL_ROW = ITEM_END_ROW + 6     # 177 税込合計
REMARK_LABEL_ROW = SUMMARY_TOTAL_ROW + 2  # 179 備考

ITEM_FONT_SIZE = 9
ITEM_ROW_HEIGHT = 27

# 1ページに載せる明細行数。ヘッダーがある1枚目は少なめ。
# 実際に載る行数より控えめにして、アプリ側が勝手に改ページを挟まないようにする。
FIRST_PAGE_ITEM_ROWS = 16
NEXT_PAGE_ITEM_ROWS = 22

MONEY_FORMAT = "#,##0"
YEN_FORMAT = '"¥"#,##0'

DISALLOWED_SHEET_NAMES = {"書き出しの概要"}
DISALLOWED_SHEET_PREFIXES = ("シート1 -", "シート1-")


# ---------------------------------------------------------------------------
# データ構造
# ---------------------------------------------------------------------------

@dataclass
class LineItem:
    name: str
    spec: str = ""
    qty: float = 0
    unit: str = ""
    unit_price: int = 0
    amount: int = 0
    remark: str = ""


@dataclass
class Category:
    name: str
    items: List[LineItem] = field(default_factory=list)

    @property
    def subtotal(self) -> int:
        return int(round(sum(i.amount for i in self.items)))


@dataclass
class Estimate:
    vendor: str            # 見積元
    customer: str          # 宛名
    project_name: str      # 工事名
    issue_date: str        # 発行日（見積日）
    valid_days: str        # 有効期限
    categories: List[Category] = field(default_factory=list)
    discount: int = 0      # 値引き（マイナス値で保持）
    tax_rate: float = 0.10

    @property
    def subtotal(self) -> int:
        return int(round(sum(c.subtotal for c in self.categories)))

    @property
    def net_total(self) -> int:
        # 税抜合計 = 小計 + 値引き（値引きは負数）
        return int(round(self.subtotal + self.discount))

    @property
    def tax(self) -> int:
        # 消費税は税抜合計に対して計算（切り捨て）
        return int(self.net_total * self.tax_rate)

    @property
    def grand_total(self) -> int:
        return int(round(self.net_total + self.tax))


# ---------------------------------------------------------------------------
# テストデータ（物置工事）
# ---------------------------------------------------------------------------

def default_test_estimate() -> Estimate:
    return Estimate(
        vendor="株式会社ティズプラス",
        customer="株式会社 貞清工務店 御中",
        project_name="物置工事",
        issue_date="2026/7/18",
        valid_days="30日",
        discount=-5192,
        tax_rate=0.10,
        categories=[
            Category("仮設工事", [
                LineItem("水盛り・やり方", "", 1, "式", 15000, 15000),
                LineItem("清掃・養生・片付け", "", 1, "式", 20000, 20000),
            ]),
            Category("基礎ブロック工事", [
                LineItem("ブロック積み", "", 9, "m", 8000, 72000),
            ]),
            Category("物置工事", [
                LineItem("FF-2618HDL-2", "定価604,000円", 1, "式", 440920, 440920),
                LineItem("組立費", "", 1, "式", 160000, 160000),
            ]),
            Category("諸経費・経費", [
                LineItem("諸経費・経費", "", 1, "式", 70000, 70000),
            ]),
        ],
    )


# ---------------------------------------------------------------------------
# フラット行 ⇔ Estimate 変換（Streamlit の手入力/補正用）
# ---------------------------------------------------------------------------

DETAIL_COLUMNS = ["工事分類", "工事項目", "仕様", "数量", "単位", "単価", "金額", "備考"]


def estimate_to_rows(estimate: Estimate) -> List[Dict]:
    rows: List[Dict] = []
    for cat in estimate.categories:
        for item in cat.items:
            rows.append({
                "工事分類": cat.name,
                "工事項目": item.name,
                "仕様": item.spec,
                "数量": item.qty,
                "単位": item.unit,
                "単価": int(round(item.unit_price or 0)),
                "金額": int(round(item.amount or 0)),
                "備考": item.remark,
            })
    return rows


def _to_number(value) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "").replace("円", "").replace("¥", "")
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def rows_to_estimate(
    rows: List[Dict],
    *,
    vendor: str,
    customer: str,
    project_name: str,
    issue_date: str,
    valid_days: str,
    discount: int,
    tax_rate: float = 0.10,
) -> Estimate:
    categories: List[Category] = []
    index: Dict[str, Category] = {}
    for row in rows:
        name = str(row.get("工事項目", "")).strip()
        cat_name = str(row.get("工事分類", "")).strip() or "工事"
        if not name:
            continue
        qty = _to_number(row.get("数量"))
        unit_price = int(round(_to_number(row.get("単価"))))
        amount_raw = _to_number(row.get("金額"))
        amount = int(round(amount_raw if amount_raw else qty * unit_price))
        item = LineItem(
            name=name,
            spec=str(row.get("仕様", "") or ""),
            qty=int(qty) if float(qty).is_integer() else qty,
            unit=str(row.get("単位", "") or ""),
            unit_price=unit_price,
            amount=amount,
            remark=str(row.get("備考", "") or ""),
        )
        if cat_name not in index:
            cat = Category(cat_name, [])
            index[cat_name] = cat
            categories.append(cat)
        index[cat_name].items.append(item)
    return Estimate(
        vendor=vendor,
        customer=customer,
        project_name=project_name,
        issue_date=issue_date,
        valid_days=valid_days,
        categories=categories,
        discount=int(round(discount)),
        tax_rate=tax_rate,
    )


# ---------------------------------------------------------------------------
# セル書き込みヘルパー
# ---------------------------------------------------------------------------

def _set(ws, row: int, col: int, value, *, number_format: Optional[str] = None):
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        return
    cell.value = value
    if number_format:
        cell.number_format = number_format


# ---------------------------------------------------------------------------
# テンプレート新規作成（openpyxl）
# ---------------------------------------------------------------------------

def build_single_sheet_template(path: Path = TEMPLATE_PATH) -> Path:
    """Excel互換の1シート見積書テンプレートを openpyxl で新規作成する。

    値は入れず、固定ラベル・レイアウト・書式のみを持つ雛形を作る。
    """
    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    bottom = Border(bottom=thin)
    header_fill = PatternFill("solid", fgColor="EFEFEF")
    base_font = Font(name=FONT_NAME, size=10)
    label_font = Font(name=FONT_NAME, size=10, bold=True)
    title_font = Font(name=FONT_NAME, size=20, bold=True)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center")

    wb = Workbook()
    ws = wb.active
    ws.title = QUOTE_SHEET

    # 既定フォント（全体の見た目を標準フォントに寄せる）
    ws.sheet_view.showGridLines = False

    # 列幅（A4縦想定）
    widths = {"A": 4.5, "B": 31, "C": 17, "D": 6.5, "E": 5, "F": 11.5, "G": 14.5, "H": 11}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    navy_font = Font(name=FONT_NAME, size=10, bold=True, color=BRAND_NAVY)
    item_font = Font(name=FONT_NAME, size=ITEM_FONT_SIZE)
    rule = Side(style="thin", color=BRAND_RULE)
    under = Border(bottom=Side(style="medium", color=BRAND_NAVY))
    box = Border(left=Side(style="medium", color=BRAND_NAVY), right=Side(style="medium", color=BRAND_NAVY),
                 top=Side(style="medium", color=BRAND_NAVY), bottom=Side(style="medium", color=BRAND_NAVY))
    navy_fill = PatternFill("solid", fgColor=BRAND_NAVY)
    tint_fill = PatternFill("solid", fgColor=BRAND_TINT)

    def merge(r1, c1, r2, c2):
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
        return ws.cell(r1, c1)

    # --- 表題バンド（紺地に白抜き） ---
    t = merge(TITLE_ROW, 1, TITLE_ROW, LAST_COL)
    t.value = "御　見　積　書"
    t.font = Font(name=FONT_NAME, size=22, bold=True, color="FFFFFF")
    t.fill = navy_fill
    t.alignment = center
    ws.row_dimensions[TITLE_ROW].height = 32
    ws.row_dimensions[1].height = 6
    ws.row_dimensions[3].height = 4

    # --- ロゴ（右上） ---
    if LOGO_PATH.exists():
        try:
            img = XLImage(str(LOGO_PATH))
            img.width, img.height = 186, 67
            ws.add_image(img, "F4")
        except Exception:
            pass

    # --- 宛名 ---
    cust = merge(CUSTOMER_ROW, COL_NO, CUSTOMER_ROW, COL_SPEC)
    cust.font = Font(name=FONT_NAME, size=15)
    cust.alignment = Alignment(horizontal="left", vertical="center")
    cust.border = under
    ws.row_dimensions[CUSTOMER_ROW].height = 24
    hon = ws.cell(CUSTOMER_ROW, COL_QTY, "御中")
    hon.font = Font(name=FONT_NAME, size=12)
    hon.alignment = Alignment(horizontal="left", vertical="center")

    g = merge(GREET_ROW, COL_NO, GREET_ROW, COL_QTY)
    g.value = "下記の通り御見積申し上げます。"
    g.font = base_font
    g.alignment = Alignment(horizontal="left", vertical="center")

    # --- 見積日 / 登録番号（右） ---
    for row, text in ((ISSUE_DATE_ROW, "見積日"), (REG_ROW, "登録番号")):
        lc = ws.cell(row, COL_PRICE, text)
        lc.font = navy_font
        lc.alignment = Alignment(horizontal="left", vertical="center")
        vc = merge(row, COL_AMOUNT, row, COL_REMARK)
        vc.font = base_font
        vc.alignment = Alignment(horizontal="left", vertical="center")
        vc.border = Border(bottom=rule)
    ws.cell(REG_ROW, COL_AMOUNT).value = COMPANY["reg"]

    # --- 御見積金額（最も目立たせる） ---
    al = merge(AMOUNT_LABEL_ROW, COL_NO, AMOUNT_END_ROW, COL_ITEM)
    al.value = "御 見 積 金 額"
    al.font = Font(name=FONT_NAME, size=13, bold=True, color="FFFFFF")
    al.fill = navy_fill
    al.alignment = center
    av = merge(AMOUNT_LABEL_ROW, COL_SPEC, AMOUNT_END_ROW, COL_UNIT)
    av.font = Font(name=FONT_NAME, size=20, bold=True, color=BRAND_NAVY)
    av.alignment = Alignment(horizontal="right", vertical="center", indent=1, shrink_to_fit=True)
    av.number_format = YEN_FORMAT
    av.border = box
    ws.row_dimensions[AMOUNT_LABEL_ROW].height = 17
    ws.row_dimensions[AMOUNT_END_ROW].height = 17
    tx = merge(AMOUNT_END_ROW + 1, COL_SPEC, AMOUNT_END_ROW + 1, COL_UNIT)
    tx.value = "（消費税込）"
    tx.font = Font(name=FONT_NAME, size=8, color="666666")
    tx.alignment = Alignment(horizontal="right", vertical="center")

    # --- 工事情報（左） ---
    for offset, text in enumerate(INFO_LABELS):
        row = INFO_START_ROW + offset
        lc = merge(row, COL_NO, row, COL_ITEM)     # A:B ラベル（右寄せで値の直前に置く）
        lc.value = text
        lc.font = navy_font
        lc.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        vc = merge(row, COL_SPEC, row, COL_QTY)    # C:D 値
        vc.font = base_font
        # 長い工事名でも右の自社情報に重ならないよう、はみ出さずに縮めて収める。
        vc.alignment = Alignment(horizontal="left", vertical="center", shrink_to_fit=True)
        vc.border = Border(bottom=rule)
        ws.row_dimensions[row].height = 18
    ws.cell(INFO_START_ROW + 3, COL_SPEC).value = "ご相談の上"
    ws.cell(INFO_START_ROW + 5, COL_SPEC).value = "中村 哲也"

    # --- 自社情報（右） ---
    company_lines = [
        (COMPANY["name"], Font(name=FONT_NAME, size=12, bold=True, color=BRAND_NAVY)),
        (f'{COMPANY["zip"]} {COMPANY["address"]}', Font(name=FONT_NAME, size=9)),
        (f'{COMPANY["tel"]}　{COMPANY["fax"]}', Font(name=FONT_NAME, size=9)),
        (COMPANY["ceo"], Font(name=FONT_NAME, size=9)),
    ]
    for offset, (text, font) in enumerate(company_lines):
        c = merge(COMPANY_ROW + offset, COL_UNIT, COMPANY_ROW + offset, COL_REMARK)
        c.value = text
        c.font = font
        c.alignment = Alignment(horizontal="left", vertical="center")

    # --- 工事内容バンド ---
    band = merge(WORK_BAND_ROW, 1, WORK_BAND_ROW, LAST_COL)
    band.value = "工 事 内 容"
    band.font = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
    band.fill = navy_fill
    band.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[WORK_BAND_ROW].height = 20

    # --- 明細テーブル ヘッダー ---
    headers = ["No.", "工事品目", "仕様", "数量", "単位", "単価", "金額", "備考"]
    for col, text in enumerate(headers, start=1):
        c = ws.cell(ITEM_HEADER_ROW, col, text)
        c.font = Font(name=FONT_NAME, size=10, bold=True, color=BRAND_NAVY)
        c.fill = tint_fill
        c.alignment = center
        c.border = Border(left=rule, right=rule,
                          top=Side(style="medium", color=BRAND_NAVY),
                          bottom=Side(style="medium", color=BRAND_NAVY))
    ws.row_dimensions[ITEM_HEADER_ROW].height = 20

    # --- 明細テーブル 本体（枠線のみ・値は空） ---
    for row in range(ITEM_START_ROW, ITEM_END_ROW + 1):
        is_last = row == ITEM_END_ROW
        for col in range(1, LAST_COL + 1):
            c = ws.cell(row, col)
            c.font = base_font
            c.border = Border(
                left=rule, right=rule, top=rule,
                bottom=Side(style="medium", color=BRAND_NAVY) if is_last else rule,
            )
            if col in (COL_NO, COL_QTY, COL_UNIT):
                c.alignment = center
            elif col in (COL_PRICE, COL_AMOUNT):
                c.alignment = Alignment(horizontal="right", vertical="center")
                c.number_format = MONEY_FORMAT
            else:
                c.alignment = left
            c.font = item_font
        # 行高を固定して改ページ位置を安定させる（自動伸長だと毎回変わり体裁が崩れる）
        ws.row_dimensions[row].height = ITEM_ROW_HEIGHT

    # --- 集計ブロック ---
    summary_rows = [
        (SUMMARY_SUBTOTAL_ROW, "小　　計", False),
        (SUMMARY_DISCOUNT_ROW, "値 引 き", False),
        (SUMMARY_NET_ROW, "税抜合計", False),
        (SUMMARY_TAX_ROW, "消 費 税", False),
        (SUMMARY_TOTAL_ROW, "税込合計", True),
    ]
    for row, text, emphasize in summary_rows:
        lc = merge(row, COL_UNIT, row, COL_PRICE)
        lc.value = text
        lc.alignment = Alignment(horizontal="center", vertical="center")
        vc = ws.cell(row, COL_AMOUNT)
        # 桁が大きくても ### にならないよう、はみ出す場合は縮めて収める
        vc.alignment = Alignment(horizontal="right", vertical="center", shrink_to_fit=True)
        vc.number_format = YEN_FORMAT if emphasize else MONEY_FORMAT
        if emphasize:
            lc.font = Font(name=FONT_NAME, size=12, bold=True, color="FFFFFF")
            lc.fill = navy_fill
            vc.font = Font(name=FONT_NAME, size=12, bold=True, color=BRAND_NAVY)
            vc.border = box
            ws.row_dimensions[row].height = 23
        else:
            lc.font = navy_font
            lc.fill = tint_fill
            lc.border = Border(left=rule, right=rule, top=rule, bottom=rule)
            vc.font = base_font
            vc.border = Border(left=rule, right=rule, top=rule, bottom=rule)
            ws.row_dimensions[row].height = 17

    # --- 備考欄 ---
    rl = ws.cell(REMARK_LABEL_ROW, COL_NO, "備考")
    rl.font = navy_font
    rl.alignment = Alignment(horizontal="left", vertical="center")
    rv = merge(REMARK_LABEL_ROW, COL_ITEM, REMARK_LABEL_ROW + 1, COL_REMARK)
    rv.font = base_font
    rv.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    rv.border = Border(left=rule, right=rule, top=rule, bottom=rule)

    note = merge(REMARK_LABEL_ROW + 3, COL_NO, REMARK_LABEL_ROW + 3, LAST_COL)
    note.value = "※本見積書の有効期限を過ぎた場合は、再度お見積りいたします。"
    note.font = Font(name=FONT_NAME, size=8, color="666666")
    note.alignment = Alignment(horizontal="left", vertical="center")

    # 余白行を詰める（A4 1枚に収まる見積を1枚のまま保つため）
    for spacer_row, spacer_height in (
        (5, 8), (9, 8), (AMOUNT_END_ROW + 1, 12), (WORK_BAND_ROW - 1, 8),
        (SUMMARY_SUBTOTAL_ROW - 1, 6), (SUMMARY_TOTAL_ROW + 1, 6),
        (REMARK_LABEL_ROW + 2, 6), (REMARK_LABEL_ROW + 3, 12),
    ):
        ws.row_dimensions[spacer_row].height = spacer_height

    # 印刷設定：A4縦・横1ページに収める
    ws.page_setup.orientation = "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    # 改ページを入れなかった＝1枚で収まる設計なので、縦も1ページに寄せる
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.6, bottom=0.5)
    ws.print_area = f"A1:H{REMARK_LABEL_ROW + 3}"

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def ensure_template(path: Path = TEMPLATE_PATH) -> Path:
    if not path.exists():
        build_single_sheet_template(path)
    return path


# ---------------------------------------------------------------------------
# 流し込み
# ---------------------------------------------------------------------------

def _copy_item_header(ws, row: int) -> None:
    """明細ヘッダー行をそのままの体裁で指定行に複製する（2枚目以降の見出し用）。"""
    for col in range(1, LAST_COL + 1):
        src = ws.cell(ITEM_HEADER_ROW, col)
        dst = ws.cell(row, col)
        if isinstance(dst, MergedCell):
            continue
        dst.value = src.value
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
    header_height = ws.row_dimensions[ITEM_HEADER_ROW].height
    if header_height:
        ws.row_dimensions[row].height = header_height


def _write_quote_sheet(ws, estimate: Estimate) -> int:
    """見積書シートへ値だけを書き込む。書き込んだ明細行数（テーブル使用行数）を返す。"""
    # テンプレート側に「御中」があるので、宛名末尾の敬称は取り除いて二重表記を防ぐ。
    customer = re.sub(r"[\s　]*(御中|様|殿)[\s　]*$", "", estimate.customer or "")
    _set(ws, CUSTOMER_ROW, COL_NO, customer)
    _set(ws, ISSUE_DATE_ROW, COL_AMOUNT, estimate.issue_date)
    _set(ws, AMOUNT_LABEL_ROW, COL_SPEC, estimate.grand_total, number_format=YEN_FORMAT)

    # 工事情報（工事名称 / 工事場所 / 工事期間 / 支払条件 / 有効期限 / 見積担当）
    _set(ws, INFO_START_ROW, COL_SPEC, estimate.project_name)
    _set(ws, INFO_START_ROW + 4, COL_SPEC, estimate.valid_days)

    cat_font = Font(name=FONT_NAME, size=10, bold=True, color=BRAND_NAVY)
    cat_fill = PatternFill("solid", fgColor=BRAND_TINT)
    sub_font = Font(name=FONT_NAME, size=10, bold=True)

    # 明細を書き出す。ページが変わる位置では見出し行を「実データとして」書き込み、
    # 明示的な改ページを入れる。Numbers は Excel の印刷タイトル行設定を無視するため、
    # 印刷設定に頼らずセルの中身として見出しを持たせないと2枚目が見出し無しになる。
    state = {"row": ITEM_START_ROW, "on_page": 0, "limit": FIRST_PAGE_ITEM_ROWS}

    def emit(fill_row) -> None:
        if state["on_page"] >= state["limit"]:
            _copy_item_header(ws, state["row"])
            ws.row_breaks.append(Break(id=state["row"] - 1))
            state["row"] += 1
            state["on_page"] = 0
            state["limit"] = NEXT_PAGE_ITEM_ROWS
        fill_row(state["row"])
        state["row"] += 1
        state["on_page"] += 1

    no = 0
    multi = len(estimate.categories) > 1
    for idx, cat in enumerate(estimate.categories, start=1):
        if multi:
            def write_category(r, idx=idx, cat=cat):
                # 分類見出し行（複数業者・複数工種のときだけ出す）
                _set(ws, r, COL_ITEM, f"{idx}. {cat.name}")
                for col in range(1, LAST_COL + 1):
                    c = ws.cell(r, col)
                    if not isinstance(c, MergedCell):
                        c.fill = cat_fill
                ws.cell(r, COL_ITEM).font = cat_font
            emit(write_category)
        for item in cat.items:
            no += 1

            def write_item(r, item=item, no=no):
                _set(ws, r, COL_NO, no)
                _set(ws, r, COL_ITEM, item.name)
                _set(ws, r, COL_SPEC, item.spec)
                _set(ws, r, COL_QTY, item.qty)
                _set(ws, r, COL_UNIT, item.unit)
                _set(ws, r, COL_PRICE, int(round(item.unit_price)), number_format=MONEY_FORMAT)
                _set(ws, r, COL_AMOUNT, int(round(item.amount)), number_format=MONEY_FORMAT)
                _set(ws, r, COL_REMARK, item.remark)
            emit(write_item)
        if multi:
            def write_subtotal(r, cat=cat):
                # 分類小計行
                _set(ws, r, COL_PRICE, "小計")
                ws.cell(r, COL_PRICE).alignment = Alignment(horizontal="right", vertical="center")
                ws.cell(r, COL_PRICE).font = sub_font
                _set(ws, r, COL_AMOUNT, cat.subtotal, number_format=MONEY_FORMAT)
                ws.cell(r, COL_AMOUNT).font = sub_font
            emit(write_subtotal)

    row = state["row"]
    used_rows = row - ITEM_START_ROW

    # 使わなかった明細行は隠して、A4 1枚に収まるようにする。
    # （行を削除すると集計セルの位置がずれて検証できなくなるため、非表示にする）
    last_used = row - 1
    if last_used >= ITEM_START_ROW:
        for col in range(1, LAST_COL + 1):
            c = ws.cell(last_used, col)
            if not isinstance(c, MergedCell):
                old = c.border
                c.border = Border(left=old.left, right=old.right, top=old.top,
                                  bottom=Side(style="medium", color=BRAND_NAVY))
    for hidden_row in range(max(row, ITEM_START_ROW), ITEM_END_ROW + 1):
        ws.row_dimensions[hidden_row].hidden = True

    # 集計ブロック
    _set(ws, SUMMARY_SUBTOTAL_ROW, COL_AMOUNT, estimate.subtotal, number_format=MONEY_FORMAT)
    _set(ws, SUMMARY_DISCOUNT_ROW, COL_AMOUNT, estimate.discount, number_format=MONEY_FORMAT)
    _set(ws, SUMMARY_NET_ROW, COL_AMOUNT, estimate.net_total, number_format=MONEY_FORMAT)
    _set(ws, SUMMARY_TAX_ROW, COL_AMOUNT, estimate.tax, number_format=MONEY_FORMAT)
    _set(ws, SUMMARY_TOTAL_ROW, COL_AMOUNT, estimate.grand_total, number_format=YEN_FORMAT)

    # 複数ページになっても体裁が崩れないようにする。
    # ・明細ヘッダー行は毎ページの先頭で繰り返す（2枚目以降が見出しなしの表にならない）
    # ・フッターに工事名とページ番号を入れて、続きものだと分かるようにする
    ws.oddFooter.left.text = f"{COMPANY['name']}　{estimate.project_name}"
    ws.oddFooter.left.size = 8
    ws.oddFooter.left.color = "808080"
    ws.oddFooter.right.text = "Page &P / &N"
    ws.oddFooter.right.size = 8
    ws.oddFooter.right.color = "808080"

    # 収まりきる小さな見積だけ1ページに強制し、長い見積は自然改ページに任せる。
    # 改ページを入れていない＝1枚で収まる設計なので、縦も1ページに寄せる
    ws.page_setup.fitToHeight = 0 if ws.row_breaks.count else 1

    return used_rows


def _write_detail_sheet(wb: Workbook, estimate: Estimate):
    """2枚目「明細データ」シートを openpyxl のセル書き込みで作成する。"""
    if DETAIL_SHEET in wb.sheetnames:
        del wb[DETAIL_SHEET]
    ws = wb.create_sheet(DETAIL_SHEET)
    ws.sheet_view.showGridLines = False

    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="EFEFEF")
    base_font = Font(name=FONT_NAME, size=10)
    label_font = Font(name=FONT_NAME, size=10, bold=True)

    headers = ["No.", "工事分類", "工事項目", "仕様", "数量", "単位", "単価", "金額", "備考"]
    widths = [4, 16, 24, 18, 6, 5, 12, 13, 16]
    for i, w in enumerate(widths):
        ws.column_dimensions[chr(ord("A") + i)].width = w
    for col, text in enumerate(headers, start=1):
        c = ws.cell(1, col, text)
        c.font = label_font
        c.fill = header_fill
        c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center")

    row = 2
    no = 0
    for cat in estimate.categories:
        for item in cat.items:
            no += 1
            values = [no, cat.name, item.name, item.spec, item.qty, item.unit,
                      int(round(item.unit_price)), int(round(item.amount)), item.remark]
            for col, value in enumerate(values, start=1):
                c = ws.cell(row, col, value)
                c.font = base_font
                c.border = border
                if col in (7, 8):
                    c.number_format = MONEY_FORMAT
            row += 1

    # 明細データもA4横1ページ幅に収め、見出し行を各ページで繰り返す。
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)
    ws.print_title_rows = "1:1"
    ws.freeze_panes = "A2"

    row += 1  # 1行空ける
    summary = [
        ("小計", estimate.subtotal),
        ("値引き", estimate.discount),
        ("税抜合計", estimate.net_total),
        ("消費税", estimate.tax),
        ("税込合計", estimate.grand_total),
    ]
    for text, value in summary:
        lc = ws.cell(row, 7, text)
        lc.font = label_font
        lc.alignment = Alignment(horizontal="right", vertical="center")
        vc = ws.cell(row, 8, value)
        vc.font = label_font if text == "税込合計" else base_font
        vc.number_format = MONEY_FORMAT
        row += 1


# ---------------------------------------------------------------------------
# 検証
# ---------------------------------------------------------------------------

@dataclass
class FillResult:
    ok: bool
    errors: List[str]
    warnings: List[str]
    data: Optional[bytes]
    file_name: str
    sheet_names: List[str]
    summary: Dict


def _validate_estimate(estimate: Estimate) -> Tuple[List[str], List[str]]:
    """データそのものの整合性を検証。(errors, warnings) を返す。

    errors はダウンロードをブロックする致命的不整合。
    warnings は出力は許すが確認を促す軽微な差異（一式行の数量×単価≠金額など）。
    """
    errors: List[str] = []
    warnings: List[str] = []
    detail_sum = int(round(sum(i.amount for c in estimate.categories for i in c.items)))
    if detail_sum != estimate.subtotal:
        errors.append(f"明細合計と小計が一致していません。（明細合計 {detail_sum:,}円 / 小計 {estimate.subtotal:,}円）")
    if estimate.net_total + estimate.tax != estimate.grand_total:
        errors.append(
            f"税抜合計＋消費税が税込合計と一致していません。"
            f"（{estimate.net_total:,} + {estimate.tax:,} ≠ {estimate.grand_total:,}）"
        )
    for idx, item in enumerate(
        (i for c in estimate.categories for i in c.items), start=1
    ):
        qty = _to_number(item.qty)
        # 一式・端数配分などで 数量×単価≠金額 は正常に起こりうるため警告に留める。
        if qty and item.unit_price and int(round(qty * item.unit_price)) != int(round(item.amount)):
            warnings.append(f"No.{idx}「{item.name}」の 数量×単価 と 金額 が一致していません（金額を優先して出力します）。")
    used = 0
    for cat in estimate.categories:
        used += 1 + len(cat.items) + 1  # 見出し + 明細 + 小計
    if used > ITEM_MAX_ROWS:
        errors.append("明細行がテンプレートの上限（1シート）を超えています。")
    return errors, warnings


def _validate_output_workbook(wb: Workbook, estimate: Estimate) -> List[str]:
    """出力ワークブックのシート構成・金額セルを検証。"""
    errors: List[str] = []
    names = wb.sheetnames

    if len(names) > 2:
        errors.append(f"出力検証に失敗しました。テンプレートが2シートを超えています。（{len(names)}シート）")
    if QUOTE_SHEET not in names:
        errors.append(f"出力検証に失敗しました。「{QUOTE_SHEET}」シートがありません。")
    if DETAIL_SHEET not in names and len(names) != 1:
        errors.append(f"出力検証に失敗しました。「{DETAIL_SHEET}」シートがありません。")
    for name in names:
        if name in DISALLOWED_SHEET_NAMES or any(name.startswith(p) for p in DISALLOWED_SHEET_PREFIXES):
            errors.append("出力検証に失敗しました。不要なNumbers分解シートが含まれています。")
            break

    if QUOTE_SHEET in names:
        q = wb[QUOTE_SHEET]
        checks = [
            (SUMMARY_SUBTOTAL_ROW, estimate.subtotal, "小計"),
            (SUMMARY_DISCOUNT_ROW, estimate.discount, "値引き"),
            (SUMMARY_NET_ROW, estimate.net_total, "税抜合計"),
            (SUMMARY_TAX_ROW, estimate.tax, "消費税"),
            (SUMMARY_TOTAL_ROW, estimate.grand_total, "税込合計"),
        ]
        for row, expected, label in checks:
            actual = q.cell(row, COL_AMOUNT).value
            if int(round(_to_number(actual))) != int(round(expected)):
                errors.append(f"出力検証に失敗しました。見積書シートの{label}が正しくありません。（{actual} ≠ {expected:,}）")

    if DETAIL_SHEET in names:
        d = wb[DETAIL_SHEET]
        detail_sum = 0
        row = 2
        while d.cell(row, 1).value is not None:
            detail_sum += int(round(_to_number(d.cell(row, 8).value)))
            row += 1
        if detail_sum != estimate.subtotal:
            errors.append(
                f"明細合計と小計が一致していません。（明細データ合計 {detail_sum:,}円 / 小計 {estimate.subtotal:,}円）"
            )
    return errors


# ---------------------------------------------------------------------------
# メイン：流し込み実行
# ---------------------------------------------------------------------------

def fill_estimate_v2(
    estimate: Estimate,
    *,
    template_path: Path = TEMPLATE_PATH,
    save_to_disk: bool = True,
) -> FillResult:
    ensure_template(template_path)

    summary = {
        "見積元": estimate.vendor,
        "宛名": estimate.customer,
        "工事名": estimate.project_name,
        "小計": estimate.subtotal,
        "値引き": estimate.discount,
        "税抜合計": estimate.net_total,
        "消費税": estimate.tax,
        "税込合計": estimate.grand_total,
    }

    # 1) データ整合性の検証（致命的な errors があれば流し込まない）
    errors, warnings = _validate_estimate(estimate)

    date_part = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe = lambda s: str(s).replace("/", "").replace("\\", "").replace(" ", "").strip()
    file_name = f"TEST_Excel互換_彩架建設見積書_{safe(estimate.project_name)}_{safe(estimate.vendor)}_{date_part}.xlsx"

    if errors:
        return FillResult(False, errors, warnings, None, file_name, [], summary)

    # 2) テンプレートを読み込み、値だけ書き込む（元テンプレートには絶対に触らない）
    template_before = template_path.read_bytes()
    wb = load_workbook(template_path)
    if wb.sheetnames[0] != QUOTE_SHEET:
        # 想定外のテンプレート構成
        return FillResult(
            False,
            [f"出力検証に失敗しました。テンプレートの1枚目が「{QUOTE_SHEET}」ではありません。"],
            [], None, file_name, list(wb.sheetnames), summary,
        )

    quote_ws = wb[QUOTE_SHEET]
    _write_quote_sheet(quote_ws, estimate)
    _write_detail_sheet(wb, estimate)

    # 3) 出力ワークブックの検証
    errors.extend(_validate_output_workbook(wb, estimate))

    # 4) 元テンプレートが更新されていないことを確認
    if template_path.read_bytes() != template_before:
        errors.append("出力検証に失敗しました。元テンプレートが更新されています。")

    if errors:
        return FillResult(False, errors, warnings, None, file_name, list(wb.sheetnames), summary)

    # 5) 別名保存 ＆ バイト取得
    output = BytesIO()
    wb.save(output)
    data = output.getvalue()
    sheet_names = list(wb.sheetnames)

    if save_to_disk:
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        (OUTPUT_DIR / file_name).write_bytes(data)

    return FillResult(True, [], warnings, data, file_name, sheet_names, summary)


# ---------------------------------------------------------------------------
# 本番OCRデータ → Estimate 変換アダプタ
# ---------------------------------------------------------------------------

def _first(row, keys, default=""):
    for k in keys:
        if k in row and str(row.get(k, "")).strip() not in ("", "nan", "None"):
            return row.get(k)
    return default


def estimate_from_production(
    detail_df,
    cost_df,
    *,
    metadata: Optional[Dict] = None,
    vendors: Optional[List[str]] = None,
    valid_days: str = "30日",
    tax_rate: float = 0.10,
) -> Estimate:
    """本番フロー（extract_data.py）の detail_df / cost_df から Estimate を組み立てる。

    - 見積元（会社）ごとに 1 カテゴリにまとめる。
    - 金額は各行の「見積金額」を優先採用（数量×単価の再計算はしない）。
      -> 小計 = 明細金額の合計 になるので検証（明細合計＝小計）を必ず満たす。
    - 値引きは本番フローに概念が無いため 0。消費税は税抜合計×tax_rate（切り捨て）。
    """
    metadata = metadata or {}

    if detail_df is None or getattr(detail_df, "empty", True):
        raise ValueError("明細データが空です。")

    all_vendors = [str(v) for v in detail_df["見積元"].dropna().astype(str).unique().tolist()]
    target_vendors = [str(v) for v in (vendors or all_vendors)]

    categories: List[Category] = []
    for vendor in target_vendors:
        group = detail_df[detail_df["見積元"].astype(str) == vendor]
        items: List[LineItem] = []
        for _, row in group.iterrows():
            row = dict(row)
            name = str(_first(row, ["品名", "工事項目", "工事品目", "名称"], "")).strip()
            amount = int(round(_to_number(_first(row, ["見積金額", "金額", "原価金額"], 0))))
            if not name and not amount:
                continue
            qty_raw = _to_number(_first(row, ["数量"], 0))
            unit_price = int(round(_to_number(_first(row, ["見積単価", "単価", "原価単価"], 0))))
            items.append(LineItem(
                name=name or "（項目名なし）",
                spec=str(_first(row, ["仕様"], "") or ""),
                qty=int(qty_raw) if float(qty_raw).is_integer() else qty_raw,
                unit=str(_first(row, ["単位"], "") or ""),
                unit_price=unit_price,
                amount=amount,
                remark=str(_first(row, ["備考"], "") or ""),
            ))
        if items:
            categories.append(Category(vendor, items))

    if not categories:
        raise ValueError("対象の見積元に明細がありません。")

    vendor_label = "／".join(c.name for c in categories)
    customer = str(_first(metadata, ["宛名", "顧客名", "取引先", "得意先"], "") or "")
    project = str(_first(metadata, ["工事名称", "工事名", "件名", "案件名"], "") or "")
    issue = str(_first(metadata, ["見積日", "発行日", "作成日"], datetime.now().strftime("%Y/%m/%d")) or "")

    return Estimate(
        vendor=vendor_label,
        customer=customer,
        project_name=project or "見積",
        issue_date=issue,
        valid_days=valid_days,
        categories=categories,
        discount=0,
        tax_rate=tax_rate,
    )


if __name__ == "__main__":
    # テンプレートを（再）生成し、テストデータで流し込みの自己確認を行う。
    tpath = build_single_sheet_template(TEMPLATE_PATH)
    print(f"[template] created: {tpath}")
    result = fill_estimate_v2(default_test_estimate())
    print(f"[fill] ok={result.ok}")
    print(f"[fill] sheets={result.sheet_names}")
    print(f"[fill] file={result.file_name}")
    print(f"[fill] summary={result.summary}")
    if result.errors:
        print("[fill] errors:")
        for e in result.errors:
            print("   -", e)
