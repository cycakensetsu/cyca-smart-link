import unittest
from io import BytesIO

import openpyxl
import pandas as pd

from estimate_pipeline import (
    DETAIL_SHEET_NAME,
    NUMBERS_OUTPUT_COLUMNS,
    OUTPUT_COLUMNS,
    QUOTE_SHEET_NAME,
    WORK_SUMMARY_SHEET_NAME,
    assign_unknown_vendors_to_pdf_vendor,
    build_intermediate_dataframe,
    build_cost_basis_dataframe,
    build_quote_summary_dataframe,
    build_vendor_work_summary_dataframe,
    build_work_summary_dataframe,
    normalize_summary_data,
    output_dataframe,
    numbers_detail_dataframe,
    parse_money,
    vendor_detail_dataframe,
    split_extraction_payload,
    split_quantity_unit,
    validate_intermediate,
)


DANJYO_RECORDS = [
    {"No": 1, "見積元": "DANJYO", "PDF小計": 1750000, "消費税": 175000, "税込合計": 1925000, "品名": "仮設足場 W600組ブラケット足場 外手摺一本 60日間", "数量": "665架㎡", "単価": "550円", "金額": "365,750円"},
    {"No": 2, "見積元": "DANJYO", "PDF小計": 1750000, "消費税": 175000, "税込合計": 1925000, "品名": "仮設足場 W600組 本足場 外・内手摺各一本 60日間", "数量": "869架㎡", "単価": "750円", "金額": "651,750円"},
    {"No": 3, "見積元": "DANJYO", "PDF小計": 1750000, "消費税": 175000, "税込合計": 1925000, "品名": "飛散防止ネット 防炎II類", "数量": "1,534架㎡", "単価": "150円", "金額": "230,100円"},
    {"No": 4, "見積元": "DANJYO", "PDF小計": 1750000, "消費税": 175000, "税込合計": 1925000, "品名": "昇降階段/2箇所", "数量": "10基", "単価": "5,000円", "金額": "50,000円"},
    {"No": 5, "見積元": "DANJYO", "PDF小計": 1750000, "消費税": 175000, "税込合計": 1925000, "品名": "壁繋ぎ", "数量": "1,534架㎡", "単価": "100円", "金額": "153,400円"},
    {"No": 6, "見積元": "DANJYO", "PDF小計": 1750000, "消費税": 175000, "税込合計": 1925000, "品名": "場内小運搬 架け・払い", "数量": "1,534架㎡", "単価": "100円", "金額": "153,400円"},
    {"No": 7, "見積元": "DANJYO", "PDF小計": 1750000, "消費税": 175000, "税込合計": 1925000, "品名": "資材運搬費 架け・払い 糸島-柳川", "数量": "1,534架㎡", "単価": "120円", "金額": "184,080円"},
    {"No": 8, "見積元": "DANJYO", "PDF小計": 1750000, "消費税": 175000, "税込合計": 1925000, "品名": "値引き", "数量": "1式", "単価": "-38,480円", "金額": "-38,480円"},
]


class EstimatePipelineTest(unittest.TestCase):
    def test_split_quantity_unit(self):
        self.assertEqual(split_quantity_unit("665架㎡"), (665.0, "架㎡"))
        self.assertEqual(split_quantity_unit("1,534架㎡"), (1534.0, "架㎡"))
        self.assertEqual(split_quantity_unit("10基"), (10.0, "基"))
        self.assertEqual(split_quantity_unit("1式"), (1.0, "式"))

    def test_parse_negative_money(self):
        self.assertEqual(parse_money("-38,480円"), -38480.0)

    def test_danjyo_records_are_preserved_and_validated(self):
        df, totals = build_intermediate_dataframe(DANJYO_RECORDS)
        self.assertEqual(len(df), 8)
        self.assertEqual(int(df["原価金額"].sum()), 1750000)
        self.assertEqual(totals["小計"], 1750000)
        self.assertEqual(totals["消費税"], 175000)
        self.assertEqual(totals["税込合計"], 1925000)
        self.assertTrue((df["品名"] == "値引き").any())
        discount = df[df["品名"] == "値引き"].iloc[0]
        self.assertEqual(int(discount["原価金額"]), -38480)
        self.assertEqual(validate_intermediate(df, totals), [])

    def test_output_columns_are_fixed(self):
        df, _ = build_intermediate_dataframe(DANJYO_RECORDS)
        self.assertEqual(output_dataframe(df).columns.tolist(), OUTPUT_COLUMNS)

    def test_numbers_detail_keeps_all_rows(self):
        df, _ = build_intermediate_dataframe(DANJYO_RECORDS)
        detail, issues = numbers_detail_dataframe(df)
        self.assertEqual(issues, [])
        self.assertEqual(len(detail), 8)
        self.assertEqual(detail.columns.tolist(), NUMBERS_OUTPUT_COLUMNS)
        self.assertEqual(detail["工事品目"].tolist(), df["品名"].tolist())
        self.assertIn("場内小運搬 架け・払い", detail["工事品目"].tolist())
        self.assertIn("資材運搬費 架け・払い 糸島-柳川", detail["工事品目"].tolist())
        self.assertIn("値引き", detail["工事品目"].tolist())
        self.assertEqual(detail["単価"].tolist(), df["見積単価"].tolist())
        self.assertEqual(detail["金額"].tolist(), df["見積金額"].tolist())
        for internal_col in ["見積元", "品名", "原価単価", "原価金額", "上乗せ額", "見積単価", "見積金額"]:
            self.assertNotIn(internal_col, detail.columns)

    def test_subtotal_mismatch_blocks_output(self):
        bad = [dict(r) for r in DANJYO_RECORDS]
        bad[0]["金額"] = "1円"
        df, totals = build_intermediate_dataframe(bad)
        issues = validate_intermediate(df, totals)
        self.assertTrue(any(issue["レベル"] == "停止" for issue in issues))

    def test_summary_payload_keeps_cover_sheet_totals(self):
        payload = {
            "page_role": "cover_summary_page",
            "summary_data": {
                "工事名称": "瀧上工業 防水改修工事",
                "工事項目": [
                    {"工事項目": "防水工事", "金額": 11381700},
                    {"工事項目": "諸経費", "金額": 65000},
                    {"工事項目": "厚生福利費", "金額": 55000},
                ],
                "小計": 11501700,
                "端数調整": -1700,
                "改小計": 11500000,
                "消費税": 1150000,
                "工事費計": 12650000,
            },
            "detail_data": [
                {"No": 1, "見積元": "瀧上工業", "品名": "防水明細A", "数量": 1, "単位": "式", "単価": 11381700, "金額": 11381700}
            ],
        }
        summaries, details = split_extraction_payload(payload)
        summary_data = normalize_summary_data(summaries)
        df, _ = build_intermediate_dataframe(details)
        summary_df, totals = build_quote_summary_dataframe(summary_data, df)

        self.assertEqual(totals["工事費計"], 12650000)
        self.assertEqual(totals["小計"], 11501700)
        values = summary_df.fillna("").astype(str).to_string()
        self.assertIn("彩架建設 見積書", values)
        self.assertIn("防水工事", values)
        self.assertIn("12650000", values)

    def test_workbook_always_starts_with_quote_work_summary_then_detail_sheet(self):
        payload = {
            "summary_data": {
                "宛名": "御中",
                "工事名称": "（元）海老津ショッピングセンター屋上防水改修工事",
                "工事場所": "海老津ショッピングセンター",
                "工事項目": [
                    {"工事項目": "防水工事", "金額": 11381700},
                    {"工事項目": "諸経費", "金額": 65000},
                    {"工事項目": "厚生福利費", "金額": 55000},
                ],
                "小計": 11501700,
                "端数調整": -1700,
                "改小計": 11500000,
                "消費税": 1150000,
                "工事費計": 12650000,
            },
            "detail_data": [
                {"No": 1, "見積元": "瀧上工業", "品名": "防水明細A", "数量": 1, "単位": "式", "単価": 11381700, "金額": 11381700}
            ],
        }
        summaries, details = split_extraction_payload(payload)
        summary_data = normalize_summary_data(summaries)
        df, _ = build_intermediate_dataframe(details)
        quote_df, totals = build_quote_summary_dataframe(summary_data, df)
        work_summary_df, work_totals = build_work_summary_dataframe(summary_data, df)
        detail_df, issues = numbers_detail_dataframe(df)
        self.assertEqual(issues, [])

        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            quote_df.to_excel(writer, index=False, header=False, sheet_name=QUOTE_SHEET_NAME)
            work_summary_df.to_excel(writer, index=False, header=False, sheet_name=WORK_SUMMARY_SHEET_NAME)
            detail_df.to_excel(writer, index=False, header=False, sheet_name=DETAIL_SHEET_NAME)

        workbook = openpyxl.load_workbook(BytesIO(output.getvalue()), read_only=True)
        self.assertEqual(workbook.sheetnames, ["見積書", "工事別まとめ", "明細データ"])
        self.assertEqual(totals["工事費計"], 12650000)
        self.assertEqual(work_totals["工事費計"], 12650000)
        quote_values = [
            cell
            for row in workbook["見積書"].iter_rows(values_only=True)
            for cell in row
            if cell not in (None, "")
        ]
        self.assertIn("彩架建設 見積書", quote_values)
        self.assertIn("（元）海老津ショッピングセンター屋上防水改修工事", quote_values)
        self.assertIn(12650000, quote_values)
        work_values = [
            cell
            for row in workbook["工事別まとめ"].iter_rows(values_only=True)
            for cell in row
            if cell not in (None, "")
        ]
        self.assertIn("防水工事", work_values)
        self.assertIn("諸経費", work_values)
        self.assertIn("厚生福利費", work_values)
        self.assertIn(12650000, work_values)

    def test_work_summary_keeps_multiple_company_summary_pages_separate(self):
        summaries = []
        details = []
        payloads = [
            (
                "瀧上工業.pdf",
                {
                    "page_role": "cover_summary_page",
                    "summary_data": {
                        "見積元": "瀧上工業",
                        "工事項目": [
                            {"工事項目": "防水工事一式", "金額": 11381700},
                            {"工事項目": "諸経費", "金額": 65000},
                            {"工事項目": "厚生福利費", "金額": 55000},
                        ],
                        "端数調整": -1700,
                    },
                },
            ),
            (
                "塗装工事.pdf",
                {
                    "page_role": "summary_page",
                    "summary_data": {
                        "見積元": "塗装会社",
                        "工事項目": [
                            {"工事項目": "外部塗装工事一式", "金額": 5000000},
                            {"工事項目": "諸経費", "金額": 300000},
                        ],
                    },
                },
            ),
            (
                "仮設工事.pdf",
                {
                    "page_role": "summary_page",
                    "summary_data": {
                        "見積元": "足場会社",
                        "工事項目": [
                            {"工事項目": "仮設工事一式", "金額": 2500000},
                        ],
                    },
                },
            ),
        ]
        for idx, (source_name, payload) in enumerate(payloads, start=1):
            page_summaries, page_details = split_extraction_payload(payload, source_name=source_name, page_number=idx)
            summaries.extend(page_summaries)
            details.extend(page_details)

        detail_records = [
            {"No": 1, "見積元": "瀧上工業", "品名": "防水明細", "数量": 1, "単位": "式", "単価": 11381700, "金額": 11381700},
            {"No": 1, "見積元": "塗装会社", "品名": "塗装明細", "数量": 1, "単位": "式", "単価": 5000000, "金額": 5000000},
            {"No": 1, "見積元": "足場会社", "品名": "足場明細", "数量": 1, "単位": "式", "単価": 2500000, "金額": 2500000},
        ]
        df, _ = build_intermediate_dataframe(detail_records)
        summary_data = normalize_summary_data(summaries)
        work_summary_df, totals = build_work_summary_dataframe(summary_data, df)

        values = work_summary_df.fillna("").astype(str).to_string()
        self.assertIn("防水工事一式", values)
        self.assertIn("外部塗装工事一式", values)
        self.assertIn("仮設工事一式", values)
        self.assertIn("瀧上工業", values)
        self.assertIn("塗装会社", values)
        self.assertIn("足場会社", values)
        self.assertEqual(totals["小計"], 19301700)
        self.assertEqual(len(work_summary_df[work_summary_df["No"].astype(str).str.strip() != ""]), 6)

    def test_vendor_summary_is_cost_basis_and_unknown_detail_stays_as_vendor_detail(self):
        summary_sources = [
            {
                "__source_name": "山田製作所.pdf",
                "__page_number": 1,
                "見積元": "株式会社 山田製作所",
                "工事項目": [
                    {"工事項目": "A部 腐食部補修", "金額": 68000},
                    {"工事項目": "B部 腐食部補強補修", "数量": 2, "単位": "箇所", "金額": 150000},
                    {"工事項目": "手摺り部分全塗装", "数量": 145, "単位": "㎡", "金額": 696000},
                ],
                "小計": 914000,
                "消費税": 91400,
                "工事費計": 1005400,
            },
            {
                "__source_name": "アキヨシ塗装.pdf",
                "__page_number": 1,
                "見積元": "アキヨシ塗装",
                "工事項目": [
                    {"工事項目": "塗装工事一式", "金額": 3800000},
                ],
                "小計": 3800000,
                "消費税": 380000,
                "工事費計": 4180000,
            },
        ]
        records = [
            {
                "__source_name": "山田製作所.pdf",
                "__page_number": 1,
                "見積元": "株式会社 山田製作所",
                "品名": "A部 腐食部補修",
                "数量": "1式",
                "単価": "68000",
                "金額": "68000",
            },
            {
                "__source_name": "山田製作所.pdf",
                "__page_number": 1,
                "見積元": "株式会社 山田製作所",
                "品名": "B部 腐食部補強補修",
                "数量": "2箇所",
                "単価": "75000",
                "金額": "150000",
            },
            {
                "__source_name": "山田製作所.pdf",
                "__page_number": 1,
                "見積元": "株式会社 山田製作所",
                "品名": "手摺り部分全塗装",
                "数量": "145㎡",
                "単価": "4800",
                "金額": "696000",
            },
            {
                "__source_name": "アキヨシ塗装.pdf",
                "__page_number": 2,
                "見積元": "不明",
                "品名": "塗装明細A",
                "数量": "1式",
                "単価": "2000000",
                "金額": "2000000",
            },
            {
                "__source_name": "アキヨシ塗装.pdf",
                "__page_number": 2,
                "見積元": "不明",
                "品名": "塗装明細B",
                "数量": "1式",
                "単価": "1578500",
                "金額": "1578500",
            },
        ]

        assigned_records, assign_debug = assign_unknown_vendors_to_pdf_vendor(records, summary_sources)
        summary_data = normalize_summary_data(summary_sources)
        detail_df, _ = build_intermediate_dataframe(assigned_records)
        cost_df, vendor_summaries = build_cost_basis_dataframe(summary_data, detail_df)
        work_df, work_totals = build_vendor_work_summary_dataframe(vendor_summaries, cost_df)
        detail_sheet_df, detail_issues = vendor_detail_dataframe(detail_df)

        self.assertTrue(any(row["assigned_from_same_pdf"] for row in assign_debug))
        self.assertNotIn("不明", detail_df["見積元"].tolist())
        self.assertEqual(int(detail_df["原価金額"].sum()), 4492500)
        self.assertEqual(int(cost_df["原価金額"].sum()), 4714000)
        self.assertEqual(cost_df["見積元"].tolist(), ["株式会社 山田製作所", "アキヨシ塗装"])
        self.assertEqual(work_totals["小計"], 4714000)
        self.assertEqual(detail_issues, [])

        work_values = work_df.fillna("").astype(str).to_string()
        detail_values = detail_sheet_df.fillna("").astype(str).to_string()
        self.assertIn("【株式会社 山田製作所 まとめ】", work_values)
        self.assertIn("【アキヨシ塗装 まとめ】", work_values)
        self.assertIn("【株式会社 山田製作所 明細】", detail_values)
        self.assertIn("【アキヨシ塗装 明細】", detail_values)

    def test_unknown_same_pdf_near_named_amount_can_still_be_removed_for_legacy_cost_flow(self):
        from estimate_pipeline import deduplicate_estimate_records
        records = [
            {
                "__source_name": "塗装見積.pdf",
                "__page_number": 2,
                "見積元": "アキヨシ塗装",
                "品名": "外壁塗装工事",
                "数量": "1式",
                "単価": "3800000",
                "金額": "3800000",
            },
            {
                "__source_name": "塗装見積.pdf",
                "__page_number": 2,
                "見積元": "不明",
                "品名": "塗装工事一式",
                "数量": "1式",
                "単価": "3578500",
                "金額": "3578500",
                "抽出元テキスト範囲": "塗装工事一式 3,578,500",
            },
        ]

        filtered, debug_rows = deduplicate_estimate_records(records)
        df, _ = build_intermediate_dataframe(filtered)

        self.assertEqual(len(filtered), 1)
        self.assertNotIn("不明", df["見積元"].tolist())
        self.assertEqual(int(df["原価金額"].sum()), 3800000)
        excluded = [row for row in debug_rows if row["重複判定で除外"]]
        self.assertEqual(len(excluded), 1)
        self.assertEqual(excluded[0]["company_name"], "不明")
        self.assertEqual(excluded[0]["duplicate_of"], "アキヨシ塗装")

    def test_summary_sheet_recalculates_after_profit(self):
        summaries = [{
            "工事項目": [
                {"工事項目": "防水工事", "金額": 11381700},
                {"工事項目": "諸経費", "金額": 65000},
                {"工事項目": "厚生福利費", "金額": 55000},
            ],
            "端数調整": -1700,
        }]
        summary_data = normalize_summary_data(summaries)
        df, _ = build_intermediate_dataframe([
            {"No": 1, "見積元": "瀧上工業", "品名": "防水明細A", "数量": 1, "単位": "式", "単価": 11381700, "金額": 11381700}
        ])
        df["上乗せ額"] = 100000
        df["見積金額"] = 11481700
        df["見積単価"] = 11481700
        _, totals = build_quote_summary_dataframe(summary_data, df)

        self.assertEqual(totals["小計"], 11601700)
        self.assertEqual(totals["改小計"], 11600000)
        self.assertEqual(totals["消費税"], 1160000)
        self.assertEqual(totals["工事費計"], 12760000)


if __name__ == "__main__":
    unittest.main()
