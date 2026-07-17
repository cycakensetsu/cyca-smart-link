from __future__ import annotations

from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell

from estimate_pipeline import normalize_text, parse_money


TEMPLATE_FILL_TEST_NAME = "テンプレート流し込みテスト"
TEMPLATE_PATH = Path(__file__).resolve().parent / "templates" / "test_cyca_estimate_template_excel_compatible.xlsx"
OUTPUT_DIR = Path(__file__).resolve().parent / "output" / "template_fill_test"

QUOTE_SHEET = "TEST_見積書"
DETAIL_SHEET = "TEST_明細データ"

QUOTE_SUMMARY_ROW = 15
QUOTE_SUMMARY_END_ROW = 34
DETAIL_START_ROW = 4
DETAIL_END_ROW = 203
DISALLOWED_SHEET_PREFIXES = ("シート1 -",)
DISALLOWED_SHEET_NAMES = {"書き出しの概要"}


def _safe_sheet(wb: Workbook, name: str):
    if name not in wb.sheetnames:
        raise ValueError(f"テスト用テンプレート内に必要なシートがありません: {name}")
    return wb[name]


def _set_value(ws, cell_ref: str, value):
    cell = ws[cell_ref]
    if isinstance(cell, MergedCell):
        return
    cell.value = value


def _set_cell(ws, row: int, col: int, value):
    cell = ws.cell(row=row, column=col)
    if not isinstance(cell, MergedCell):
        cell.value = value


def _clear_range_values(ws, start_row: int, end_row: int, start_col: int = 1, end_col: int = 8):
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            _set_cell(ws, row, col, None)


def _row_from_series(row: pd.Series, amount_col: str = "見積金額", unit_col: str = "見積単価") -> Dict:
    return {
        "工事項目": normalize_text(row.get("品名", row.get("工事項目", ""))),
        "仕様": normalize_text(row.get("仕様", "")),
        "数量": row.get("数量", ""),
        "単位": normalize_text(row.get("単位", "")),
        "単価": int(round(parse_money(row.get(unit_col)) or 0)),
        "金額": int(round(parse_money(row.get(amount_col)) or 0)),
        "備考": normalize_text(row.get("備考", "")),
    }


def _detail_rows_for_vendor(detail_df: pd.DataFrame, vendor_name: str) -> List[Dict]:
    if detail_df is None or detail_df.empty or "見積元" not in detail_df.columns:
        return []
    rows = []
    vendor_df = detail_df[detail_df["見積元"].astype(str) == str(vendor_name)]
    for _, row in vendor_df.iterrows():
        item = _row_from_series(row)
        if item["工事項目"] or item["金額"]:
            rows.append(item)
    return rows


def _write_quote_sheet(ws, metadata: Dict, vendor_name: str, subtotal: int, tax: int, total: int):
    _set_value(ws, "G4", metadata.get("見積日", datetime.now().strftime("%Y年%m月%d日")))
    _set_value(ws, "B4", metadata.get("宛名", metadata.get("顧客名", "")))
    _set_value(ws, "B6", metadata.get("工事名称", metadata.get("件名", "")))
    _set_value(ws, "B7", metadata.get("工事場所", ""))
    _set_value(ws, "B8", metadata.get("工事期間", ""))
    _set_value(ws, "B9", metadata.get("支払条件", "ご相談の上"))
    _set_value(ws, "B10", metadata.get("有効期限", "お打ち合わせの上"))
    _set_value(ws, "B11", metadata.get("見積担当", "中村哲也"))

    _set_value(ws, "G6", subtotal)
    _set_value(ws, "G7", tax)
    _set_value(ws, "G8", total)

    _clear_range_values(ws, QUOTE_SUMMARY_ROW, QUOTE_SUMMARY_END_ROW)
    _set_cell(ws, QUOTE_SUMMARY_ROW, 2, f"{vendor_name} 工事一式")
    _set_cell(ws, QUOTE_SUMMARY_ROW, 4, 1)
    _set_cell(ws, QUOTE_SUMMARY_ROW, 5, "式")
    _set_cell(ws, QUOTE_SUMMARY_ROW, 6, subtotal)
    _set_cell(ws, QUOTE_SUMMARY_ROW, 7, subtotal)
    _set_cell(ws, QUOTE_SUMMARY_ROW, 8, "TEST直接流し込み")

    _set_value(ws, "G36", subtotal)
    _set_value(ws, "G37", tax)
    _set_value(ws, "G38", total)


def _write_detail_sheet(ws, detail_rows: List[Dict], subtotal: int, tax: int, total: int):
    _clear_range_values(ws, DETAIL_START_ROW, DETAIL_END_ROW)
    for offset, row in enumerate(detail_rows[: DETAIL_END_ROW - DETAIL_START_ROW + 1]):
        excel_row = DETAIL_START_ROW + offset
        values = [
            offset + 1,
            row.get("工事項目", ""),
            row.get("仕様", ""),
            row.get("数量", ""),
            row.get("単位", ""),
            row.get("単価", ""),
            row.get("金額", ""),
            row.get("備考", ""),
        ]
        for col, value in enumerate(values, start=1):
            _set_cell(ws, excel_row, col, value)

    total_row = DETAIL_START_ROW + len(detail_rows[: DETAIL_END_ROW - DETAIL_START_ROW + 1]) + 1
    if total_row + 2 <= DETAIL_END_ROW:
        _set_cell(ws, total_row, 6, "小計")
        _set_cell(ws, total_row, 7, subtotal)
        _set_cell(ws, total_row + 1, 6, "消費税")
        _set_cell(ws, total_row + 1, 7, tax)
        _set_cell(ws, total_row + 2, 6, "税込合計")
        _set_cell(ws, total_row + 2, 7, total)


def validate_template_fill_test(detail_rows: List[Dict], original_subtotal: int, markup_amount: int, subtotal: int) -> List[str]:
    issues: List[str] = []
    detail_sum = int(round(sum(parse_money(row.get("金額")) or 0 for row in detail_rows)))
    if detail_sum != subtotal:
        issues.append(f"明細金額の合計({detail_sum:,}円)と小計({subtotal:,}円)が一致しません。")
    if original_subtotal + markup_amount != subtotal:
        issues.append(f"原価({original_subtotal:,}円)+上乗せ({markup_amount:,}円)が上乗せ後小計({subtotal:,}円)と一致しません。")
    for idx, row in enumerate(detail_rows, start=1):
        qty = parse_money(row.get("数量")) or 0
        unit_price = parse_money(row.get("単価")) or 0
        amount = parse_money(row.get("金額")) or 0
        if qty and unit_price and int(round(qty * unit_price)) != int(round(amount)):
            issues.append(f"No.{idx} 数量×単価と金額が一致しません。")
    if len(detail_rows) > DETAIL_END_ROW - DETAIL_START_ROW + 1:
        issues.append("明細行がテスト用テンプレートの上限を超えています。")
    return issues


def _validate_output_workbook(wb: Workbook, subtotal: int, detail_rows: List[Dict]) -> List[str]:
    issues: List[str] = []
    if len(wb.sheetnames) > 2:
        issues.append(f"出力シート数が2枚を超えています: {len(wb.sheetnames)}枚")
    for required in [QUOTE_SHEET, DETAIL_SHEET]:
        if required not in wb.sheetnames:
            issues.append(f"必要なシートがありません: {required}")
    for name in wb.sheetnames:
        if name in DISALLOWED_SHEET_NAMES or any(name.startswith(prefix) for prefix in DISALLOWED_SHEET_PREFIXES):
            issues.append(f"不要なNumbers分解シートが残っています: {name}")
    if QUOTE_SHEET in wb.sheetnames:
        quote_ws = wb[QUOTE_SHEET]
        if int(round(parse_money(quote_ws["G6"].value) or 0)) != subtotal:
            issues.append("見積書シートの小計が正しくありません。")
    if DETAIL_SHEET in wb.sheetnames:
        detail_ws = wb[DETAIL_SHEET]
        detail_sum = 0
        for row in range(DETAIL_START_ROW, DETAIL_START_ROW + len(detail_rows)):
            detail_sum += int(round(parse_money(detail_ws.cell(row, 7).value) or 0))
        if detail_sum != subtotal:
            issues.append(f"明細データシートの明細合計({detail_sum:,}円)と小計({subtotal:,}円)が一致しません。")
    return issues


def build_template_fill_test_workbook(
    *,
    detail_df: pd.DataFrame,
    cost_df: pd.DataFrame,
    vendor_name: Optional[str] = None,
    metadata: Optional[Dict] = None,
    template_path: Path = TEMPLATE_PATH,
) -> Tuple[bytes, str, List[str]]:
    """Load the Excel-compatible test template and fill values into fixed cells only."""
    if not template_path.exists():
        raise FileNotFoundError(f"テスト用Excel互換テンプレートが見つかりません: {template_path}")
    if cost_df is None or cost_df.empty:
        raise ValueError("集計対象データがありません。")

    metadata = metadata or {}
    vendor_names = [normalize_text(v) for v in cost_df["見積元"].dropna().astype(str).unique().tolist()]
    selected_vendor = normalize_text(vendor_name) or (vendor_names[0] if vendor_names else "")
    if not selected_vendor:
        raise ValueError("見積元が取得できません。")

    cost_vendor_df = cost_df[cost_df["見積元"].astype(str) == str(selected_vendor)]
    detail_vendor_df = (
        detail_df[detail_df["見積元"].astype(str) == str(selected_vendor)]
        if detail_df is not None and not detail_df.empty and "見積元" in detail_df.columns
        else pd.DataFrame()
    )
    original_subtotal = int(round(pd.to_numeric(detail_vendor_df.get("原価金額", pd.Series(dtype=float)), errors="coerce").fillna(0).sum()))
    subtotal = int(round(pd.to_numeric(cost_vendor_df["見積金額"], errors="coerce").fillna(0).sum()))
    markup_amount = subtotal - original_subtotal
    tax = int(round(subtotal * 0.10))
    total = subtotal + tax

    detail_rows = _detail_rows_for_vendor(detail_df, selected_vendor)
    issues = validate_template_fill_test(detail_rows, original_subtotal, markup_amount, subtotal)

    template_before = template_path.read_bytes()
    wb = load_workbook(template_path)
    quote_ws = _safe_sheet(wb, QUOTE_SHEET)
    detail_ws = _safe_sheet(wb, DETAIL_SHEET)

    _write_quote_sheet(quote_ws, metadata, selected_vendor, subtotal, tax, total)
    _write_detail_sheet(detail_ws, detail_rows, subtotal, tax, total)
    issues.extend(_validate_output_workbook(wb, subtotal, detail_rows))

    if template_path.read_bytes() != template_before:
        issues.append("元テンプレートが変更されています。")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    date_part = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_vendor = selected_vendor.replace("/", "").replace("\\", "")
    file_name = f"TEST_彩架建設見積テンプレート流し込み_{safe_vendor}_{date_part}.xlsx"
    output_path = OUTPUT_DIR / file_name
    output = BytesIO()
    wb.save(output)
    data = output.getvalue()
    output_path.write_bytes(data)
    return data, file_name, issues
